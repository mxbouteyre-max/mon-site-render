"""
app.py — Serveur Flask pour Render

Logique :
  - Lister les scripts disponibles (code_extraction/*.py)
  - Lancer un ou plusieurs scripts sélectionnés (en arrière-plan)
  - Suivre l'état des jobs (running / finished / failed)
  - Une fois terminés, fusionner tous les résultats (csv/xlsx/tsv)
    produits dans code_extraction/ en un seul fichier .xlsx téléchargeable
"""

import os, sys, glob, io, re, subprocess, threading, unicodedata
from datetime import datetime
from flask import Flask, jsonify, request, send_from_directory, send_file

app = Flask(__name__)

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
SCRIPTS_DIR = os.path.join(BASE_DIR, "code_extraction")
JOBS = {}
JOBS_LOCK = threading.Lock()
MAX_RUNNING = 2

DATA_EXTENSIONS = (".csv", ".xlsx", ".tsv")

# ── référentiel des noms de colonnes (basé sur bdd_compilee, feuille Feuil1) ─
# Ordre = ordre de sortie souhaité dans le fichier final. Toute colonne d'un
# fichier source qui correspond (exactement ou par alias) à une de ces
# colonnes sera renommée pour matcher exactement ce nom.
REFERENCE_COLUMNS = [
    "enseigne", "nom", "url", "adresse", "cp", "ville", "departement",
    "nom_departement", "region", "pays", "telephone", "latitude", "longitude",
    "email", "fichier_source", "google_maps", "phone_raw", "siret", "siren",
    "url_doctolib", "zone", "zone recherchée", "fax", "siret_api",
    "verification", "adresse_sirene", "nom_sirene", "score_adresse",
    "tel_clean", "doublon", "dédoublonage",
]

# Alias connus -> colonne de référence. Les clés sont déjà normalisées
# (cf. _normalize_colname) : minuscules, sans accents, sans espaces/tirets.
COLUMN_ALIASES = {
    # enseigne
    "enseigne": "enseigne", "marque": "enseigne", "brand": "enseigne",
    "chaine": "enseigne", "reseau": "enseigne",
    # nom (du magasin/boutique)
    "nom": "nom", "nommagasin": "nom", "nomboutique": "nom",
    "nomenseigne": "nom", "name": "nom", "store": "nom", "magasin": "nom",
    "boutique": "nom", "titre": "nom",
    # url
    "url": "url", "lien": "url", "link": "url", "siteweb": "url",
    "site": "url", "pageurl": "url",
    # adresse
    "adresse": "adresse", "address": "adresse", "adressecomplete": "adresse",
    "rue": "adresse", "voie": "adresse",
    # code postal
    "cp": "cp", "codepostal": "cp", "zipcode": "cp", "postalcode": "cp",
    "zip": "cp",
    # ville
    "ville": "ville", "city": "ville", "commune": "ville",
    # departement
    "departement": "departement", "dept": "departement", "dpt": "departement",
    "codedepartement": "departement",
    # nom_departement
    "nomdepartement": "nom_departement", "departementnom": "nom_departement",
    # region
    "region": "region",
    # pays
    "pays": "pays", "country": "pays",
    # telephone
    "telephone": "telephone", "tel": "telephone", "phone": "telephone",
    "numerotelephone": "telephone", "telfixe": "telephone",
    "telephonefixe": "telephone",
    # latitude / longitude
    "latitude": "latitude", "lat": "latitude",
    "longitude": "longitude", "lng": "longitude", "lon": "longitude",
    "long": "longitude",
    # email
    "email": "email", "mail": "email", "courriel": "email", "emailaddress": "email",
    # fichier_source
    "fichiersource": "fichier_source", "source": "fichier_source",
    "sourcefile": "fichier_source",
    # google_maps
    "googlemaps": "google_maps", "lienmaps": "google_maps",
    "googlemapsurl": "google_maps", "mapsurl": "google_maps",
    # phone_raw
    "phoneraw": "phone_raw", "telbrut": "phone_raw", "telephonebrut": "phone_raw",
    "telraw": "phone_raw",
    # siret
    "siret": "siret",
    # siren
    "siren": "siren",
    # url_doctolib
    "urldoctolib": "url_doctolib", "doctolib": "url_doctolib",
    "liendoctolib": "url_doctolib",
    # zone
    "zone": "zone",
    # zone recherchée
    "zonerecherchee": "zone recherchée", "zonerecherche": "zone recherchée",
    # fax
    "fax": "fax", "numerofax": "fax",
    # siret_api
    "siretapi": "siret_api",
    # verification
    "verification": "verification", "verif": "verification",
    "verifie": "verification",
    # adresse_sirene
    "adressesirene": "adresse_sirene",
    # nom_sirene
    "nomsirene": "nom_sirene",
    # score_adresse
    "scoreadresse": "score_adresse",
    # tel_clean
    "telclean": "tel_clean", "telephoneclean": "tel_clean",
    "telephonenettoye": "tel_clean", "telnettoye": "tel_clean",
    # doublon
    "doublon": "doublon", "duplicate": "doublon",
    # dédoublonage
    "dedoublonage": "dédoublonage",
}


def _normalize_colname(name):
    """minuscule, sans accents, sans espaces/tirets/underscores superflus."""
    name = str(name).strip().lower()
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    name = re.sub(r"[\s\-_]+", "", name)
    return name


def harmonize_columns(df):
    """
    Renomme les colonnes de df pour matcher le référentiel REFERENCE_COLUMNS,
    via correspondance exacte (normalisée) puis via la table d'alias.
    Les colonnes sans correspondance gardent leur nom d'origine.
    En cas de collision (plusieurs colonnes du fichier source pointant vers
    le même nom cible), seule la première est renommée ; les suivantes
    gardent leur nom d'origine pour ne pas créer de doublon ambigu.
    """
    ref_by_norm = {_normalize_colname(c): c for c in REFERENCE_COLUMNS}
    rename_map = {}
    used_targets = set()

    for col in df.columns:
        norm = _normalize_colname(col)

        if norm in ref_by_norm:
            target = ref_by_norm[norm]
        elif norm in COLUMN_ALIASES:
            target = COLUMN_ALIASES[norm]
        else:
            target = None

        if target and target != col and target not in used_targets:
            rename_map[col] = target
            used_targets.add(target)
        elif target:
            used_targets.add(col if target in used_targets else target)

    if rename_map:
        df = df.rename(columns=rename_map)
    return df


# ── lecture CSV robuste ───────────────────────────────────────────
def _read_csv_robust(filepath):
    """Essaie plusieurs séparateurs et encodages jusqu'à trouver le bon."""
    import pandas as pd

    encodings  = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    separators = [None, ";", ",", "\t", "|"]  # None = sniff automatique

    last_error = None
    for enc in encodings:
        for sep in separators:
            try:
                kwargs = dict(encoding=enc, on_bad_lines="skip")
                if sep is None:
                    kwargs["sep"] = None
                    kwargs["engine"] = "python"
                else:
                    kwargs["sep"] = sep
                df = pd.read_csv(filepath, **kwargs)
                if len(df.columns) > 1 or sep is None:
                    return df
            except Exception as e:
                last_error = e
                continue

    raise ValueError(f"Impossible de lire le CSV ({last_error})")


# ── exécution d'un script ─────────────────────────────────────────
def execute_script(script_name, script_path):
    with JOBS_LOCK:
        JOBS[script_name] = {
            "status": "running",
            "start_time": datetime.now().isoformat()
        }

    try:
        process = subprocess.Popen(
            [sys.executable, script_path],
            cwd=SCRIPTS_DIR,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        with JOBS_LOCK:
            JOBS[script_name]["pid"] = process.pid

        stdout, stderr = process.communicate()

        with JOBS_LOCK:
            JOBS[script_name].update({
                "status": "finished" if process.returncode == 0 else "failed",
                "returncode": process.returncode,
                "stdout": stdout[-10000:] if stdout else "",
                "stderr": stderr[-10000:] if stderr else "",
                "end_time": datetime.now().isoformat()
            })

    except Exception as e:
        with JOBS_LOCK:
            JOBS[script_name] = {
                "status": "failed",
                "error": str(e),
                "end_time": datetime.now().isoformat()
            }


# ── pages statiques ───────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/information.html")
def information():
    return send_from_directory(".", "information.html")

@app.route("/video_utilisation_site.mp4")
def video_tuto():
    return send_from_directory(".", "video_utilisation_site.mp4")

@app.route("/excel-merger.html")
def excel_merger():
    return send_from_directory(".", "excel-merger.html")

@app.route("/fusion")
def fusion():
    return send_from_directory(".", "fusion-colonnes.html")

@app.route("/linkedIn")
def linkedIn():
    return send_from_directory(".", "parser_linkedIn.html")


# ── liste des scripts disponibles ─────────────────────────────────
@app.route("/api/scripts")
def api_scripts():
    if not os.path.isdir(SCRIPTS_DIR):
        os.makedirs(SCRIPTS_DIR)
    files = sorted([
        os.path.basename(p)
        for p in glob.glob(os.path.join(SCRIPTS_DIR, "*.py"))
    ])
    return jsonify({"scripts": files})


# ── lancer un ou plusieurs scripts ────────────────────────────────
@app.route("/api/run", methods=["POST"])
def api_run():
    data = request.get_json(force=True)

    # accepte soit "script": "x.py" soit "scripts": ["x.py", "y.py"]
    names = data.get("scripts")
    if names is None:
        single = data.get("script", "")
        names = [single] if single else []

    if not names:
        return jsonify({"ok": False, "error": "Aucun script sélectionné"}), 400

    results = []

    for name in names:
        if ".." in name or "/" in name or "\\" in name:
            results.append({"script": name, "ok": False, "error": "Nom invalide"})
            continue

        path = os.path.join(SCRIPTS_DIR, name)

        if not os.path.isfile(path):
            results.append({"script": name, "ok": False, "error": f"Script introuvable : {name}"})
            continue

        with JOBS_LOCK:
            already_running = JOBS.get(name, {}).get("status") == "running"
            running_count = sum(1 for job in JOBS.values() if job.get("status") == "running")

        if already_running:
            results.append({"script": name, "ok": False, "error": "Script déjà en cours"})
            continue

        if running_count >= MAX_RUNNING:
            results.append({"script": name, "ok": False, "error": f"Maximum {MAX_RUNNING} scripts simultanés"})
            continue

        thread = threading.Thread(
            target=execute_script,
            args=(name, path),
            daemon=True
        )
        thread.start()

        results.append({"script": name, "ok": True, "message": f"{name} lancé en arrière-plan"})

    return jsonify({"ok": True, "results": results})


# ── état des jobs ──────────────────────────────────────────────────
@app.route("/api/jobs")
def api_jobs():
    with JOBS_LOCK:
        return jsonify(JOBS)


# ── liste des fichiers de résultats produits ───────────────────────
@app.route("/api/outputs")
def api_outputs():
    files = sorted([
        os.path.basename(f)
        for f in glob.glob(os.path.join(SCRIPTS_DIR, "*"))
        if os.path.basename(f).endswith(DATA_EXTENSIONS)
    ])
    return jsonify({"files": files})


# ── télécharger un fichier individuel ──────────────────────────────
@app.route("/api/download/<filename>")
def api_download(filename):
    if ".." in filename or "/" in filename or "\\" in filename:
        return jsonify({"error": "Nom invalide"}), 400
    path = os.path.join(SCRIPTS_DIR, filename)
    if not os.path.isfile(path):
        return jsonify({"error": "Fichier introuvable"}), 404
    return send_file(path, as_attachment=True, download_name=filename)


# ── télécharger tous les résultats fusionnés en un xlsx ─────────────
# ?mode=single  → tout dans une seule feuille (concaténation)
# ?mode=multi   → une feuille par fichier (défaut)
@app.route("/api/download_all")
def api_download_all():
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "pandas/openpyxl non installés"}), 500

    mode = request.args.get("mode", "multi")  # "single" ou "multi"

    files = sorted([
        f for f in glob.glob(os.path.join(SCRIPTS_DIR, "*"))
        if os.path.basename(f).endswith(DATA_EXTENSIONS)
    ])

    if not files:
        return jsonify({"error": "Aucun fichier de données trouvé"}), 404

    wb = Workbook()
    wb.remove(wb.active)
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center")

    def _write_sheet(ws, df):
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) else 0
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
        ws.freeze_panes = "A2"

    def _read_file(filepath):
        if filepath.endswith(".xlsx"):
            return pd.read_excel(filepath)
        elif filepath.endswith(".tsv"):
            return pd.read_csv(filepath, sep="\t", encoding="utf-8-sig", on_bad_lines="skip")
        else:
            return _read_csv_robust(filepath)

    if mode == "single":
        # ── Mode feuille unique : concat de tous les fichiers ──────────────
        dfs = []
        for filepath in files:
            try:
                df = _read_file(filepath)
                df = harmonize_columns(df)
                # Ajoute une colonne fichier_source si elle n'existe pas
                if "fichier_source" not in df.columns:
                    df.insert(0, "fichier_source", os.path.basename(filepath))
                dfs.append(df)
            except Exception as e:
                # Crée une ligne d'erreur pour ne pas perdre la trace
                dfs.append(pd.DataFrame([{"fichier_source": os.path.basename(filepath),
                                           "erreur": str(e)}]))

        if not dfs:
            return jsonify({"error": "Aucune donnée lisible"}), 404

        df_all = pd.concat(dfs, ignore_index=True, sort=False)

        # Réordonne selon REFERENCE_COLUMNS pour les colonnes connues
        ref_cols = [c for c in REFERENCE_COLUMNS if c in df_all.columns]
        other_cols = [c for c in df_all.columns if c not in ref_cols]
        df_all = df_all[ref_cols + other_cols]

        ws = wb.create_sheet("Données compilées")
        _write_sheet(ws, df_all)

    else:
        # ── Mode multi-feuilles (comportement original) ─────────────────────
        for filepath in files:
            sheet_name = os.path.basename(filepath)
            for ext in DATA_EXTENSIONS:
                sheet_name = sheet_name.replace(ext, "")
            sheet_name = sheet_name[:31]

            try:
                df = _read_file(filepath)
            except Exception as e:
                ws = wb.create_sheet(sheet_name)
                ws["A1"] = f"Erreur de lecture : {e}"
                continue

            df = harmonize_columns(df)
            ws = wb.create_sheet(sheet_name)
            _write_sheet(ws, df)

    if not wb.sheetnames:
        return jsonify({"error": "Aucune donnée lisible"}), 404

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_str = datetime.now().strftime("%Y-%m-%d")
    suffix = "_compile" if mode == "single" else ""
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"collecte_{date_str}{suffix}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8765))
    app.run(host="0.0.0.0", port=port, debug=False)

# ── suppression des fichiers de résultats après téléchargement ────────────────
@app.route("/api/clear_outputs", methods=["POST"])
def api_clear_outputs():
    deleted = []
    errors  = []
    files = glob.glob(os.path.join(SCRIPTS_DIR, "*"))
    for f in files:
        if os.path.basename(f).endswith(DATA_EXTENSIONS):
            try:
                os.remove(f)
                deleted.append(os.path.basename(f))
            except Exception as e:
                errors.append({"file": os.path.basename(f), "error": str(e)})
    return jsonify({"ok": True, "deleted": deleted, "errors": errors})